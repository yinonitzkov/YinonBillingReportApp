from sqlalchemy import create_engine
import streamlit as st
from operator import index
from os import read
from sys import implementation
from textwrap import wrap
#import pyodbc
import pandas as pd
import os.path
from os import path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from streamlit_option_menu import option_menu


SERVER = 'sql'
DATABASE = 'SCEXPERT'
DRIVER ='SQL Server'
USERNAME = 'sa'
PASSWORD = 'ga@123'
DATEBASE_CONNECTION = f'mssql://{USERNAME}:{PASSWORD}@{SERVER}/{DATABASE}?driver={DRIVER}'


engine = create_engine(f'mssql+pyodbc://{SERVER}/{DATABASE}?driver=SQL+Server+Native+Client+11.0')
connection = engine.connect()
#data = pd.read_sql_query("select top 1 * from sku", connection)

st.set_page_config('YinonRepBillingApp')

def local_css(file_name):
    with open(file_name) as f:
        st.markdown('<style>{}</style>'.format(f.read()), unsafe_allow_html=True)
local_css("style.css")

def Choose_Fill(Value):
    if Value == 'Header_fill':
        return Header_fill
    if Value == 'TotalCharge_fill':
        return TotalCharge_fill

def Choose_Font(Value):
    if Value == 'Header_font':
        return Header_font

def ExcelDesign(Full_path):
        #שליפת שמות לשוניות האקסל
                        Excel_sheets_names = pd.ExcelFile(Full_path)
                        Excel_sheets_names = Excel_sheets_names.sheet_names  # see all sheet names

                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    
                    # ריצה על כל הגליונות 
                        for SheetName in Excel_sheets_names:
                            # ריצה על כל תאי הכותרת לעדכון פונט וצבע תא
                            for HeaderCell in my_header:          
                                ExcelFile[SheetName][HeaderCell].fill = Header_fill
                                ExcelFile[SheetName][HeaderCell].font = Header_font
                            
                            # עדכון על העמודות של האקסל לרוחב מותאם
                            for SheetName in Excel_sheets_names:
                             for col in ExcelFile[SheetName].columns:
                                max_length = 0
                                column = col[0].column_letter # Get the column name
                                for cell in col:
                                    try: # Necessary to avoid error on empty cells
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = (max_length + 2) * 1.05
                                ExcelFile[SheetName].column_dimensions[column].width = adjusted_width
                        
                        ExcelFile.save(Full_path) 

ch1, ch2, ch3 = st.columns([3,1,1])
with ch3:                                    
    st.image('GLZ_LOGO.png' , width=130, ) 
with ch1:                                    
    new_title = '<p style="font-family:sans-serif; color:black; font-size: 35px; text-align: center ; "> yinon\'s billing report app</p>'
    st.markdown(new_title, unsafe_allow_html=True)

with st.sidebar:
    choose = option_menu("App Menu", ["Excel Report", "Report Design", "Wms changes", "Compare Report"],
                         icons=['file-earmark-excel', 'layers-half', 'align-center', 'bar-chart-line-fill'],
                         menu_icon="app-indicator", default_index=0,
                         styles={
        "container": {"padding": "5!important", "background-color": "#fafafa", "font-family":"Sans-serif"},
        "icon": {"color": "#f69000", "font-size": "25px"}, 
        "nav-link": {"font-size": "16px", "text-align": "left", "margin":"0px", "--hover-color": "#eee"},
        #"nav-link-selected": {"background-color": "#02ab21"},#87CEEB
        "nav-link-selected": {"background-color": "#87CEEB"},
    }
    )
if choose == "Excel Report":
    PageName = '<p style="font-family:sans-serif; color:black; font-size: 20px; background: #F0F2F6; text-align: center; ">Excel Report </p>'
    st.markdown(PageName, unsafe_allow_html=True)



    #connection = pyodbc.connect("Driver={SQL Server}; Server=sql; Database=SCEXPERT")
    #cursor = connection.cursor()
    consignee = pd.read_sql_query("select consignee from consignee UNION select 'KRAVITZ' as consignee UNION select 'ORSHAR' as consignee UNION select 'GOLF_CD' as consignee \
                                    UNION select 'ESHED' as consignee UNION select 'KNS' as consignee UNION select 'LOGISTEAM' as consignee ", connection)
    YearOptions = pd.read_sql_query("select distinct cast(year(BILLFROMDATE) as char) as BILLFROMDATE from BILLINGCHARGESDETAIL", connection)['BILLFROMDATE'].values.tolist()
    Password= '123456'
    Error_path = ''


    my_Column_header = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG']   
    my_header = ['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1','S1','T1','U1','V1','W1','X1','Y1','Z1','AA1','AB1','AC1','AD1','AE1','AF1','AG1']
    # עדכון כותרות שיקבלו הדגשה וצבע רקע
    Header_fill = PatternFill(start_color='F0FFFF', end_color='F0FFFF', fill_type='solid')
    Header_font = Font(bold=True) 
    TotalCharge_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

   ################אלקטרה################################## 
    def ELECTRA_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):   
        
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}'", connection)
                        loads_billing.drop(['SKUDESC','SKUGROUP'], axis=1, inplace=True)
                        Receipt_Billing = pd.read_sql_query(f"select * from repProformaInPerUnits where CHARGEID='{ChargeID}'", connection)
                        Receipt_Billing.drop(['skudesc','SKUGROUP','BILLTOTALCHARGE','BILLBASIS','UNITS','BILLTOTALCHARGELINE','TRANSACTIONTYPE'], axis=1, inplace=True)
                        HafatzaKod11_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Sochen='11' and Mispar_Sapak='300' and Teur_Murzar<>'הובלה מיוחדת' ", connection)
                        HafatzaMishtachim_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Teur_Murzar='משטח' and Mispar_Sapak='300' ", connection)
                        HafatzaHovalaMeyuchedet_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Teur_Murzar='הובלה מיוחדת' and Mispar_Sapak='300' ", connection)
                        HafatzaBeitLakoach_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Sochen IN ('6') and Mispar_Sapak='300' ", connection)
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Sochen<>'6' and Mispar_Sapak='300' and Sochen<>'11' and Teur_Murzar not in ('הובלה מיוחדת','משטח') ", connection)
                        HafatzaChiuvimMeyuchadim_Billing = pd.read_sql_query(f"select N'' as [תאריך] , N'' AS [תיאור החיוב                         .] , N'' AS [סכום לחיוב] , N'' AS [שם המאשר] , N'' AS [הערות                                                                                       .]", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)
                        Hafatza_Billing = pd.read_sql_query(f"select * from vPivotCheshbonSapakimMegicBill where  year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and CONSIGNEE='ELECTRA' ", connection)
                                            
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Receipt_Billing.to_excel(writer, 'כניסה', index=False, freeze_panes=[1,0],)
                            HafatzaKod11_Billing.to_excel(writer, 'הפצה קוד 11', index=False, freeze_panes=[1,0],)
                            HafatzaMishtachim_Billing.to_excel(writer, 'משטחים', index=False, freeze_panes=[1,0],)
                            HafatzaHovalaMeyuchedet_Billing.to_excel(writer, 'הובלות מיוחדות', index=False, freeze_panes=[1,0],)
                            HafatzaBeitLakoach_Billing.to_excel(writer, 'הפצה לבית לקוח' , index=False, freeze_panes=[1,0],)
                            HafatzaReport_Billing.to_excel(writer, 'דוח הפצה', index=False, freeze_panes=[1,0],)
                            HafatzaChiuvimMeyuchadim_Billing.to_excel(writer, 'חיובים מיוחדים', index=False, freeze_panes=[1,0],)
                            Hafatza_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)
                 
                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)   
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)  

                       

                    ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                        #Charge_cell = ['B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15','B17']
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'                     

                        ExcelFile.save(Full_path) 

################מחסני חשמל################################## 
    def MCHASHMAL_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):   
        
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}'", connection)
                        loads_billing.drop(['SKUDESC','SKUGROUP'], axis=1, inplace=True)
                        Receipt_Billing = pd.read_sql_query(f"select * from repProformaInPerUnits where CHARGEID='{ChargeID}'", connection)
                        Receipt_Billing.drop(['skudesc','SKUGROUP','BILLTOTALCHARGE','BILLBASIS','UNITS','BILLTOTALCHARGELINE','TRANSACTIONTYPE'], axis=1, inplace=True)
                        HafatzaMishtachim_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Teur_Murzar='משטח' and Mispar_Sapak='301' ", connection)
                        HafatzaHovalaMeyuchedet_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Teur_Murzar='הובלה מיוחדת' and Mispar_Sapak='301' ", connection)
                        HafatzaBeitLakoach_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Sochen='2' and Mispar_Sapak='301' ", connection)
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Sochen<>'6' and Mispar_Sapak='301' and Sochen<>'2' and Teur_Murzar not in ('הובלה מיוחדת','משטח') ", connection)
                        HafatzaChiuvimMeyuchadim_Billing = pd.read_sql_query(f"select N'' as [תאריך] , N'' AS [תיאור החיוב                         .] , N'' AS [סכום לחיוב] , N'' AS [שם המאשר] , N'' AS [הערות                                                                                       .]", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)
                        Hafatza_Billing = pd.read_sql_query(f"select * from vPivotCheshbonSapakimMegicBill where  year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and CONSIGNEE='MCHASHMAL' ", connection)
                                            
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Receipt_Billing.to_excel(writer, 'כניסה', index=False, freeze_panes=[1,0],)
                            HafatzaMishtachim_Billing.to_excel(writer, 'משטחים', index=False, freeze_panes=[1,0],)
                            HafatzaHovalaMeyuchedet_Billing.to_excel(writer, 'הובלות מיוחדות', index=False, freeze_panes=[1,0],)
                            HafatzaBeitLakoach_Billing.to_excel(writer, 'הפצה לבית לקוח' , index=False, freeze_panes=[1,0],)
                            HafatzaReport_Billing.to_excel(writer, 'דוח הפצה', index=False, freeze_panes=[1,0],)
                            HafatzaChiuvimMeyuchadim_Billing.to_excel(writer, 'חיובים מיוחדים', index=False, freeze_panes=[1,0],)
                            Hafatza_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)
                 
                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)   
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)  

                       

                    ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                        #Charge_cell = ['B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15','B17']
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'                     

                        ExcelFile.save(Full_path) 

################תן אלקטריק################################## 
    def TENELECTRIK_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):   
        
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        Receipt_Billing = pd.read_sql_query(f"select * from repProformaInPerUnits where CHARGEID='{ChargeID}'", connection)
                        Receipt_Billing.drop(['skudesc','SKUGROUP','BILLTOTALCHARGE','BILLBASIS','UNITS','BILLTOTALCHARGELINE','TRANSACTIONTYPE'], axis=1, inplace=True)                        
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}'", connection)
                        loads_billing.drop(['SKUDESC','SKUGROUP'], axis=1, inplace=True)
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak='310' ", connection)
                        HafatzaMischarit_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Sochen='1' and Mispar_Sapak='310' ", connection)
                        HafatzaBeitLakoach_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Sochen='2' and Mispar_Sapak='310' ", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)
                                            
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Receipt_Billing.to_excel(writer, 'כניסה', index=False, freeze_panes=[1,0],)
                            HafatzaReport_Billing.to_excel(writer, 'דוח הפצה', index=False, freeze_panes=[1,0],)
                            HafatzaMischarit_Billing.to_excel(writer, 'הפצה מסחרית', index=False, freeze_panes=[1,0],)
                            HafatzaBeitLakoach_Billing.to_excel(writer, 'הפצה לבית לקוח' , index=False, freeze_panes=[1,0],)
                            
                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)   
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)  

                       

                    ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                        #Charge_cell = ['B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15','B17']
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'                     

                        ExcelFile.save(Full_path) 

################שרות הוגן################################# 
    def SHEROTHOGEN_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):   
        
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        Receipt_Billing = pd.read_sql_query(f"select * from repProformaInPerUnits where CHARGEID='{ChargeID}'", connection)
                        Receipt_Billing.drop(['skudesc','SKUGROUP','BILLTOTALCHARGE','BILLBASIS','UNITS','BILLTOTALCHARGELINE','TRANSACTIONTYPE'], axis=1, inplace=True)                        
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}'", connection)
                        loads_billing.drop(['SKUDESC','SKUGROUP'], axis=1, inplace=True)
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak='320' ", connection)
                        HafatzaMischarit_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Sochen='1' and Mispar_Sapak='320' ", connection)
                        HafatzaBeitLakoach_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Sochen='2' and Mispar_Sapak='320' ", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)
                                            
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Receipt_Billing.to_excel(writer, 'כניסה', index=False, freeze_panes=[1,0],)
                            HafatzaReport_Billing.to_excel(writer, 'דוח הפצה', index=False, freeze_panes=[1,0],)
                            HafatzaMischarit_Billing.to_excel(writer, 'הפצה מסחרית', index=False, freeze_panes=[1,0],)
                            HafatzaBeitLakoach_Billing.to_excel(writer, 'הפצה לבית לקוח' , index=False, freeze_panes=[1,0],)
                            
                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)   
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)  

                       

                    ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                        #Charge_cell = ['B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15','B17']
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'                     

                        ExcelFile.save(Full_path) 

    ################אפרודיטה################################## 
    def AFRODITA_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):   
        
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}'", connection)
                        loads_billing.drop(['SKUDESC','SKUGROUP'], axis=1, inplace=True)
                        Mecholot_Billing = pd.read_sql_query(f"select * from vBillMECHOLOT where CHARGEID='{ChargeID}' and CONSIGNEE='AFRODITA'", connection)              
                        Receipt_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTNAME, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                    where CHARGEID='{ChargeID}' and isnull(FIELDVALUE,'')<>'MECHOLOT'  ", connection)
                        Likut_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and  CHARGEDESCRIPTION like '%ליקוט%' ", connection)
                        Hafatza_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak='66' ", connection)
                        Asortiment_Billing = pd.read_sql_query(f"select * from repProformaWcbOXESLTR where CHARGEID='{ChargeID}' and AGREEMENTLINE in ('6','7','8','9','10','11','12','13','14')", connection)
                        TeudotLMA_Billing = pd.read_sql_query(f"select '' ", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)
                                                                  
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Mecholot_Billing.to_excel(writer, 'מכולות', index=False, freeze_panes=[1,0],)
                            Receipt_Billing.to_excel(writer, 'קליטה', index=False, freeze_panes=[1,0],)
                            Likut_Billing.to_excel(writer, 'ליקוט', index=False, freeze_panes=[1,0],)
                            Hafatza_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)
                            Asortiment_Billing.to_excel(writer, 'ערך מוסף אסורטימנטים LTR', index=False, freeze_panes=[1,0],)
                            TeudotLMA_Billing.to_excel(writer, 'LMAתעודות', index=False, freeze_panes=[1,0],)
                                                                  
                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)   
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)                         

                    ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                        #Charge_cell = ['B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15','B17']
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'                     

                        ExcelFile.save(Full_path) 

################בוניטה################################## 
    def BONITA_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):   
        
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}'", connection)
                        loads_billing.drop(['SKUDESC','SKUGROUP'], axis=1, inplace=True)
                        Mecholot_Billing = pd.read_sql_query(f"select * from vBillMECHOLOT where CHARGEID='{ChargeID}' and CONSIGNEE='BONITA'", connection)              
                        Receipt_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTNAME, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                    where CHARGEID='{ChargeID}' and isnull(FIELDVALUE,'')<>'MECHOLOT'  ", connection)
                        Likut_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and  CHARGEDESCRIPTION like '%ליקוט%' ", connection)
                        Hafatza_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak='92' ", connection)
                        Asortiment_Billing = pd.read_sql_query(f"select * from repProformaWcbOXESLTR where CHARGEID='{ChargeID}' and AGREEMENTLINE in ('6','7','8','9','10','11','12','13','14')", connection)
                        TeudotLMA_Billing = pd.read_sql_query(f"select '' ", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)
                                                                  
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Mecholot_Billing.to_excel(writer, 'מכולות', index=False, freeze_panes=[1,0],)
                            Receipt_Billing.to_excel(writer, 'קליטה', index=False, freeze_panes=[1,0],)
                            Likut_Billing.to_excel(writer, 'ליקוט', index=False, freeze_panes=[1,0],)
                            Hafatza_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)
                            Asortiment_Billing.to_excel(writer, 'ערך מוסף אסורטימנטים LTR', index=False, freeze_panes=[1,0],)
                            TeudotLMA_Billing.to_excel(writer, 'LMAתעודות', index=False, freeze_panes=[1,0],)
                                                                  
                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)   
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)                         

                    ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                        #Charge_cell = ['B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15','B17']
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'                     

                        ExcelFile.save(Full_path) 
  
    ################תדיראן גרופ################################## 
    def TADIRAN_GROUP_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):   
        
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}'", connection)
                        loads_billing.drop(['SKUDESC','SKUGROUP'], axis=1, inplace=True)
                        Mecholot_Billing = pd.read_sql_query(f"select * from vBillMECHOLOT where CHARGEID='{ChargeID}' and CONSIGNEE='TADIRANG'", connection)              
                        Receipt_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTNAME, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                    where CHARGEID='{ChargeID}' and isnull(FIELDVALUE,'')<>'MECHOLOT'  ", connection)
                        Likut_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and  CHARGEDESCRIPTION like '%ליקוט%' ", connection)
                        Hafatza_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak='35' ", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)
                                                                  
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Mecholot_Billing.to_excel(writer, 'מכולות', index=False, freeze_panes=[1,0],)
                            Receipt_Billing.to_excel(writer, 'קליטה', index=False, freeze_panes=[1,0],)
                            Likut_Billing.to_excel(writer, 'ליקוט', index=False, freeze_panes=[1,0],)
                            Hafatza_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)
                                                                  
                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)   
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)                         

                    ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 20

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                        #Charge_cell = ['B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15','B17']
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'                     

                        ExcelFile.save(Full_path) 

    ################פלקסטרוניקס################################## 
    def FLEX_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):   
        
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}'", connection)
                        loads_billing.drop(['SKUDESC','SKUGROUP'], axis=1, inplace=True)
                        Receipt_HovalaAtzmit_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTNAME, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                    where CHARGEID='{ChargeID}' and CHARGELINE='1'  ", connection)
                        Receipt_Yevu_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTNAME, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                    where CHARGEID='{ChargeID}' and CHARGELINE='2'   ", connection)
                        Likut_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and  CHARGEDESCRIPTION like '%ליקוט%' ", connection)
                        Hafatza_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak='84' ", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)
                                                                  
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Receipt_Yevu_Billing.to_excel(writer, 'כניסות - יבוא', index=False, freeze_panes=[1,0],)
                            Receipt_HovalaAtzmit_Billing.to_excel(writer, 'כניסות - הובלה עצמית חיצוני', index=False, freeze_panes=[1,0],)
                            Likut_Billing.to_excel(writer, 'ליקוט', index=False, freeze_panes=[1,0],)
                            Hafatza_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)
                                                                  
                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)   
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)                         

                    ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 20

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                        #Charge_cell = ['B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15','B17']
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'                     

                        ExcelFile.save(Full_path) 

    ################מאניה גינס################################## 
    def MANIA_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):   
        
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}'", connection)
                        loads_billing.drop(['SKUDESC','SKUGROUP'], axis=1, inplace=True)
                        Mecholot_Billing = pd.read_sql_query(f"select * from vBillMECHOLOT where CHARGEID='{ChargeID}' and CONSIGNEE='BONITA'", connection)              
                        Receipt_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTNAME, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                    where CHARGEID='{ChargeID}' and isnull(FIELDVALUE,'')<>'MECHOLOT'  ", connection)
                        Likut_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and  AGREEMENTLINE='2' ", connection)
                        Hafatza_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak='92' ", connection)
                        
                        TeudotLMA_Billing = pd.read_sql_query(f"select '' ", connection)
                        Asortiment_Kvutza_Mida_Billing = pd.read_sql_query(f"select '' as DATE_SRIKA, '' as Mispar_Mishtach, '' as Mispar_PL, '' as Sug_Mishtach, '' as Sug_Carton, '' as Sku, '' as Units, '' as Mechiron, '' as TotalPrice ", connection)
                        Asortiment_Makat = pd.read_sql_query(f"select '' as DATE_SRIKA, '' as Mispar_Mishtach, '' as Mispar_PL, '' as Sug_Mishtach, '' as Sug_Carton, '' as Sku, '' as Units, '' as Mechiron, '' as TotalPrice ", connection)
                        Asortiment_Sdarot = pd.read_sql_query(f"select '' as DATE_SRIKA, '' as Mispar_Mishtach, '' as Mispar_PL, '' as Sug_Mishtach, '' as Sug_Carton, '' as Sku, '' as Units, '' as Mechiron, '' as TotalPrice ", connection)
                        Asortiment_Parit_Nailon = pd.read_sql_query(f"select '' as DATE_SRIKA, '' as Mispar_Mishtach, '' as Mispar_PL, '' as Sug_Mishtach, '' as Sug_Carton, '' as Sku, '' as Units, '' as Mechiron, '' as TotalPrice ", connection)
                        Asortiment_Nailon_ChomerGelem = pd.read_sql_query(f"select '' as DATE_SRIKA, '' as Mispar_Mishtach, '' as Mispar_PL, '' as Sug_Mishtach, '' as Sug_Carton, '' as Sku, '' as Units, '' as Mechiron, '' as TotalPrice ", connection)
                        Asortiment_Hadbakat_Barkod = pd.read_sql_query(f"select '' as DATE_SRIKA, '' as Mispar_Mishtach, '' as Mispar_PL, '' as Sug_Mishtach, '' as Sug_Carton, '' as Sku, '' as Units, '' as Mechiron, '' as TotalPrice ", connection)
                        Mishtach_Etz = pd.read_sql_query(f"select '' as DATE_SRIKA, '' as Mispar_Mishtach, '' as Mispar_PL, '' as Sug_Mishtach, '' as Sug_Carton, '' as Sku, '' as Units, '' as Mechiron, '' as TotalPrice ", connection)
                        PeulotMeyuchadot= pd.read_sql_query(f"select '' as DATE_SRIKA, '' as Mispar_Mishtach, '' as Mispar_PL, '' as Sug_Mishtach, '' as Sug_Carton, '' as Sku, '' as Units, '' as Mechiron, '' as TotalPrice ", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)
                                                                  
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Mecholot_Billing.to_excel(writer, 'מכולות', index=False, freeze_panes=[1,0],)
                            Receipt_Billing.to_excel(writer, 'קליטה', index=False, freeze_panes=[1,0],)
                            Likut_Billing.to_excel(writer, 'ליקוט', index=False, freeze_panes=[1,0],)
                            Hafatza_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)
                            TeudotLMA_Billing.to_excel(writer, 'LMAתעודות', index=False, freeze_panes=[1,0],)
                            Asortiment_Kvutza_Mida_Billing.to_excel(writer, 'מיון לקבוצה מידה', index=False, freeze_panes=[1,0],)
                            Asortiment_Makat.to_excel(writer, 'מיון לרמת מקט', index=False, freeze_panes=[1,0],)
                            Asortiment_Sdarot.to_excel(writer, 'בניית סדרות', index=False, freeze_panes=[1,0],)
                            Asortiment_Parit_Nailon.to_excel(writer, 'הכנסת פריט לניילון', index=False, freeze_panes=[1,0],)
                            Asortiment_Nailon_ChomerGelem.to_excel(writer, 'ניילון חומר גלם', index=False, freeze_panes=[1,0],)
                            Asortiment_Hadbakat_Barkod.to_excel(writer, 'הדבקת ברקוד', index=False, freeze_panes=[1,0],)
                            Mishtach_Etz.to_excel(writer, 'משטחי עץ', index=False, freeze_panes=[1,0],)
                            PeulotMeyuchadot.to_excel(writer, 'פעולות מיוחדות', index=False, freeze_panes=[1,0],)
                                                                  
                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)   
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)                         

                    ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                        #Charge_cell = ['B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15','B17']
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'                     

                        ExcelFile.save(Full_path) 

    ################בורגר ראנצ################################## 
    def BUR_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}'", connection)
                        loads_billing.drop(['SKUDESC','SKUGROUP'], axis=1, inplace=True)
                        Mecholot_Billing = pd.read_sql_query(f"select * from vBillMECHOLOT where CHARGEID='{ChargeID}' and CONSIGNEE='BUR'", connection)              
                        KlitatMasait_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                         where CHARGEID='{ChargeID}' and AGREEMENTLINE in ('1','10') ", connection)
                        Likutim_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' ", connection)
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak='52'  ", connection)          
                        MishtacheiEtzIn_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                         where CHARGEID='{ChargeID}' and AGREEMENTLINE in ('12') ", connection)
                        MishtacheiEtzOut_Billing = pd.read_sql_query(f"select * from repProformaPalltes where CHARGEID='{ChargeID}' ", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)                  
                        
                        # #רשימת השליפות ללשוניות
                        # Excel_sheets_query_names = ['loads_billing','Mecholot_Billing','KlitatMasait_Billing','Likutim_Billing'
                        #                             ,'HafatzaReport_Billing','MishtacheiEtzIn_Billing','MishtacheiEtzOut_Billing','Rikuz_Billing']

                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Mecholot_Billing.to_excel(writer, 'קליטת מכולות', index=False, freeze_panes=[1,0],)
                            KlitatMasait_Billing.to_excel(writer, 'קליטת משאית', index=False, freeze_panes=[1,0],)
                            Likutim_Billing.to_excel(writer, 'ליקוטים', index=False, freeze_panes=[1,0],)
                            HafatzaReport_Billing.to_excel(writer, 'דוח הפצה', index=False, freeze_panes=[1,0],)
                            MishtacheiEtzIn_Billing.to_excel(writer, 'משטחי עץ-כניסה', index=False, freeze_panes=[1,0],)
                            MishtacheiEtzOut_Billing.to_excel(writer, 'משטחי עץ-יציאה', index=False, freeze_panes=[1,0],)                       

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    
                        # בניית תוכן של לשונית ריכוז 
                        ExcelFile['ריכוז'] ['A3'] = 'סוג החיוב'
                        ExcelFile['ריכוז'] ['B3'] = 'מחיר'
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14
                        ExcelFile['ריכוז'] ['A3'].fill = Header_fill
                        ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        ExcelFile['ריכוז'] ['A1'].font = Header_font
                        
                        ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        ExcelFile['ריכוז'] ['A5'] = 'קליטת מכולות'
                        ExcelFile['ריכוז'] ['B5'] = '=SUM(\'קליטת מכולות\'!M:M)'
                        ExcelFile['ריכוז'] ['A6'] = 'קליטת משאיות'
                        ExcelFile['ריכוז'] ['B6'] = '=SUM(\'קליטת משאית\'!L:L)'
                        ExcelFile['ריכוז'] ['A7'] = 'ליקוטים'
                        ExcelFile['ריכוז'] ['B7'] = '=SUM(\'ליקוטים\'!P:P)'
                        ExcelFile['ריכוז'] ['A8'] = 'הובלה'
                        ExcelFile['ריכוז'] ['B8'] = '=SUM(\'דוח הפצה\'!AA:AA)'
                        ExcelFile['ריכוז'] ['A9'] = 'ביטוח'
                        ExcelFile['ריכוז'] ['B9'] = '=1500000*0.11%'
                        ExcelFile['ריכוז'] ['A10'] = 'ערך מוסף'
                        ExcelFile['ריכוז'] ['B10'] = '=0'
                        ExcelFile['ריכוז'] ['A11'] = 'משטחי עץ-כניסה'
                        ExcelFile['ריכוז'] ['B11'] = '=SUM(\'משטחי עץ-כניסה\'!L:L)'
                        ExcelFile['ריכוז'] ['A12'] = 'משטחי עץ-יציאה'
                        ExcelFile['ריכוז'] ['B12'] = '=SUM(\'משטחי עץ-יציאה\'!L:L)'
                        ExcelFile['ריכוז'] ['A13'] = 'הפרשי קיזוז'
                        ExcelFile['ריכוז'] ['B13'] = '=5000*1'
                        ExcelFile['ריכוז'] ['D13'] = 'מתוך'
                        ExcelFile['ריכוז'] ['E13'] = '36'
                        ExcelFile['ריכוז'] ['A16'] = 'סה"כ לחיוב'
                        ExcelFile['ריכוז'] ['B16'] = '=SUM(B4:B14)'
                        ExcelFile['ריכוז'] ['A16'].fill = TotalCharge_fill
                        ExcelFile['ריכוז'] ['B16'].fill = TotalCharge_fill

                        
                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ['B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B16']
                        for CellFormat in Charge_cell:
                            ExcelFile['ריכוז'] [CellFormat].number_format = u'#,##0 ₪'

                        ExcelFile.save(Full_path)  

################(BFL) פודאפיל ################################## 
    def BFL_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}' and AGREEMENTLINE in ('3')", connection)
                        Mecholot_Billing = pd.read_sql_query(f"select * from vBillMECHOLOT where CHARGEID='{ChargeID}' and CONSIGNEE='BIAPAL' AND AGREEMENTLINE IN ('203','204','205','206','207','217') ", connection)              
                        SapakimMekomieim_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and isnull(FIELDVALUE,'')<>'MECHOLOT' and ISNULL(TRANSPORTTYPE,'')='WTW' ", connection)
                        MishtacheiYevu_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                    where CHARGEID='{ChargeID}' and isnull(FIELDVALUE,'')<>'MECHOLOT' and ISNULL(TRANSPORTTYPE,'')='PurchOrder' ", connection)
                        Hachzarot_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                    where 1=2 ", connection)
                        LikutB2B_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and isnull(ORDERTYPE,'מסחרי')='מסחרי' and CHARGEDESCRIPTION like '%ליקוט%' ", connection)
                        LikutB2C_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and ORDERTYPE='ליקוט' and CHARGEDESCRIPTION like '%ליקוט%' and AGREEMENTLINE='6' ", connection)
                        ArizatOnLine_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and ORDERTYPE='ליקוט' and CHARGEDESCRIPTION like '%אריז%' ", connection)
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak='89'  ", connection)          
                        HafatzaReport_Billing["Qty"] = pd.to_numeric(HafatzaReport_Billing["Qty"])
                        TeumeiAspakaB2B_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and ORDERTYPE='מסחרי' and CHARGEDESCRIPTION like '%תיאו%' ", connection)             
                        HafatzaBulimB2B_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and AGREEMENTLINE='9' ", connection)
                        Govaina_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and CHARGEDESCRIPTION like '%גוביינ%' ", connection)
                        HaamasaAtzmit_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where 1=2 ", connection)
                        HaavaraBeinChanuiot_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where 1=2 ", connection)                  
                        ChiyuvMishtacheiEtzBaldarut_Billing = pd.read_sql_query(f"select * from ProformaBaldarutPallets where CHARGEID='{ChargeID}' and LINE in ('208') ", connection)
                        ChiyuvMishtacheiEtzHafatza_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaValueAdded where CHARGEID='{ChargeID}' and AGREEMENTLINE='220' ", connection)
                        ZikuyMishtachim_Billing = pd.read_sql_query(f"select * from ProformaBaldarutPallets where CHARGEID='{ChargeID}' and CONSIGNEE='BIAPAL' AND LINE IN ('216') ", connection)              
                        ErechMusaf_Billing = pd.read_sql_query(f"select * from ProformaSpecialBilling where CHARGEID='{ChargeID}' ", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)                  
                        
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Mecholot_Billing.to_excel(writer, 'קליטת מכולות', index=False, freeze_panes=[1,0],)
                            SapakimMekomieim_Billing.to_excel(writer, 'קליטת ספקים מקומיים', index=False, freeze_panes=[1,0],)
                            MishtacheiYevu_Billing.to_excel(writer, 'קליטת משטחי יבוא', index=False, freeze_panes=[1,0],)
                            Hachzarot_Billing.to_excel(writer, 'קליטת החזרות', index=False, freeze_panes=[1,0],)       
                            LikutB2B_Billing.to_excel(writer, 'ליקוט B2B', index=False, freeze_panes=[1,0],)
                            LikutB2C_Billing.to_excel(writer, 'ליקוט B2C', index=False, freeze_panes=[1,0],)
                            ArizatOnLine_Billing.to_excel(writer, 'אריזת הזמנות און ליין', index=False, freeze_panes=[1,0],)
                            HafatzaReport_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)
                            TeumeiAspakaB2B_Billing.to_excel(writer, 'תיאומי אספקה B2B', index=False, freeze_panes=[1,0],)
                            HafatzaBulimB2B_Billing.to_excel(writer, 'הפצה B2B בולים', index=False, freeze_panes=[1,0],)
                            Govaina_Billing.to_excel(writer, 'הפצה B2B גוביינא', index=False, freeze_panes=[1,0],)
                            HaamasaAtzmit_Billing.to_excel(writer, 'העמסה עצמית', index=False, freeze_panes=[1,0],)
                            HaavaraBeinChanuiot_Billing.to_excel(writer, 'העברות בין חנויות והחזרות', index=False, freeze_panes=[1,0],)                       
                            ChiyuvMishtacheiEtzBaldarut_Billing.to_excel(writer, 'משטחי עץ און ליין', index=False, freeze_panes=[1,0],)
                            ChiyuvMishtacheiEtzHafatza_Billing.to_excel(writer, 'חיוב משטחי עץ הפצה', index=False, freeze_panes=[1,0],)
                            ZikuyMishtachim_Billing.to_excel(writer, 'זיכוי משטחים', index=False, freeze_panes=[1,0],)
                            ErechMusaf_Billing.to_excel(writer, 'עבודות ערך מוסף', index=False, freeze_panes=[1,0],)

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    

                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14
                        ExcelFile['ריכוז'] .column_dimensions['D'].width = 15
                        ExcelFile['ריכוז'] .column_dimensions['E'].width = 15
                        ExcelFile['ריכוז'] .column_dimensions['F'].width = 15

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                        #Charge_cell = ['B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B14','B15','B17']
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      
                        # # בניית תוכן של לשונית ריכוז 
                        # ExcelFile['ריכוז'] ['A3'] = 'סוג החיוב'
                        # ExcelFile['ריכוז'] ['B3'] = 'מחיר'
                        # ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        # ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        # ExcelFile['ריכוז'] .column_dimensions['B'].width = 14
                        # ExcelFile['ריכוז'] ['A3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!U:U)'
                        # ExcelFile['ריכוז'] ['A5'] = 'קליטת מכולות'
                        # ExcelFile['ריכוז'] ['B5'] = '=SUM(\'קליטת מכולות\'!M:M)'
                        # ExcelFile['ריכוז'] ['A6'] = 'קליטת משאיות'
                        # ExcelFile['ריכוז'] ['B6'] = '=SUM(\'קליטת משאית\'!L:L)'
                        # ExcelFile['ריכוז'] ['A7'] = 'ליקוטים'
                        # ExcelFile['ריכוז'] ['B7'] = '=SUM(\'ליקוטים\'!P:P)'
                        # ExcelFile['ריכוז'] ['A8'] = 'הובלה'
                        # ExcelFile['ריכוז'] ['B8'] = '=SUM(\'דוח הפצה\'!AA:AA)'
                        # ExcelFile['ריכוז'] ['A9'] = 'ביטוח'
                        # ExcelFile['ריכוז'] ['B9'] = '=1500000*0.11%'
                        # ExcelFile['ריכוז'] ['A10'] = 'ערך מוסף'
                        # ExcelFile['ריכוז'] ['B10'] = '=0'
                        # ExcelFile['ריכוז'] ['A11'] = 'משטחי עץ-כניסה'
                        # ExcelFile['ריכוז'] ['B11'] = '=SUM(\'משטחי עץ-כניסה\'!J:J)*-16'
                        # ExcelFile['ריכוז'] ['A12'] = 'משטחי עץ-יציאה'
                        # ExcelFile['ריכוז'] ['B12'] = '=SUM(\'משטחי עץ-יציאה\'!J:J)*16'
                        # ExcelFile['ריכוז'] ['A13'] = 'הפרשי קיזוז'
                        # ExcelFile['ריכוז'] ['B13'] = '=5000*1'
                        # ExcelFile['ריכוז'] ['D13'] = 'מתוך'
                        # ExcelFile['ריכוז'] ['E13'] = '36'
                        # ExcelFile['ריכוז'] ['A16'] = 'סה"כ לחיוב'
                        # ExcelFile['ריכוז'] ['B16'] = '=SUM(B4:B14)'
                        # ExcelFile['ריכוז'] ['A16'].fill = TotalCharge_fill
                        # ExcelFile['ריכוז'] ['B16'].fill = TotalCharge_fill

                        
                        # # להפוך את כל התאים של מחיר לפורמט שקל
                        # Charge_cell = ['B4','B5','B6','B7','B8','B9','B10','B11','B12','B13','B16']
                        # for CellFormat in Charge_cell:
                        #     ExcelFile['ריכוז'] [CellFormat].number_format = u'#,##0 ₪'

                        ExcelFile.save(Full_path)

################ כיתן  ################################## 
    def KITAN_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select CHARGEID, CHARGELINE, AGREEMENTLINE, CHARGETEXT, BILLFROMDATE, BILLTODATE, BILLTOTAL, UNITS from vBillingChargesByRunIDDetail where CHARGEID={ChargeID} and AGREEMENTLINE in ('6','14')", connection)
                        Likut_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and CHARGEDESCRIPTION like '%ליקוט%' ", connection)
                        Mecholot_Billing = pd.read_sql_query(f"select * from vBillMECHOLOT where CHARGEID='{ChargeID}' and AGREEMENTLINE='28'", connection)              
                        Klita_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and isnull(FIELDVALUE,'')<>'MECHOLOT' and ISNULL(TRANSPORTTYPE,'')<>'RETURN' AND AGREEMENTLINE IN ('17','27')", connection)
                        KlitaChazarot_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and AGREEMENTLINE IN ('8','9','24') ", connection)
                        MishtacheiEtz_ZikuyHafatza_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaValueAdded where CHARGEID='{ChargeID}' and AGREEMENTLINE='40' ", connection)
                        MishtacheiEtz_ChiyuvHafatza_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaValueAdded where CHARGEID='{ChargeID}' and AGREEMENTLINE='38' ", connection)
                        #MishtacheiEtzOut_Billing = pd.read_sql_query(f"select * from CheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak in ('53','54') and Teur_Murzar='משטח'  ", connection)
                        MishtacheiBaldar_Chiyuv_Billing = pd.read_sql_query(f"select * from ProformaBaldarutPallets where CHARGEID='{ChargeID}' and LINE='36' ", connection)
                        MishtacheiBaldar_Zikuy_Billing = pd.read_sql_query(f"select * from ProformaBaldarutPallets where CHARGEID='{ChargeID}' and LINE='39' ", connection)
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak in ('53','54')  ", connection)          
                        ErechMusaf_Billing = pd.read_sql_query(f"select * from ProformaSpecialBilling where CHARGEID='{ChargeID}' ", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)                  
                        MishtacheiEtz_ChiyuvHafatza_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaValueAdded where CHARGEID='{ChargeID}' and AGREEMENTLINE='38' ", connection)
                        PickPivotReport_Billing = pd.read_sql_query(f"SELECT CHARGEDESCRIPTION, OBJECTUNITS AS Document_Units, PRICEPERUNIT, VALUE from repProformaDetailed where CHARGEID='{ChargeID}' and CHARGEDESCRIPTION like '%ליקוט%' ", connection)          
                        PickPivotReport_Billing = PickPivotReport_Billing.pivot_table(index='CHARGEDESCRIPTION', values=['Document_Units','PRICEPERUNIT','VALUE'], aggfunc=[sum],margins=True, margins_name='סה"כ')
                        
                        

                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Likut_Billing.to_excel(writer, 'ליקוטים', index=False, freeze_panes=[1,0],) 
                            PickPivotReport_Billing.to_excel(writer, 'PIVOT ליקוטים', freeze_panes=[1,0],startrow=1)
                            Mecholot_Billing.to_excel(writer, 'מכולות', index=False, freeze_panes=[1,0],)
                            Klita_Billing.to_excel(writer, 'קליטה', index=False, freeze_panes=[1,0],)
                            KlitaChazarot_Billing.to_excel(writer, 'קליטת חזרות', index=False, freeze_panes=[1,0],)
                            MishtacheiEtz_ChiyuvHafatza_Billing.to_excel(writer, 'משטחי עץ - חיוב הפצה', index=False, freeze_panes=[1,0],)
                            MishtacheiEtz_ZikuyHafatza_Billing.to_excel(writer, 'משטחי עץ - זיכוי הפצה', index=False, freeze_panes=[1,0],)
                            #MishtacheiEtzOut_Billing.to_excel(writer, 'משטחי עץ-יציאה', index=False, freeze_panes=[1,0],)
                            MishtacheiBaldar_Chiyuv_Billing.to_excel(writer, 'חיוב משטחי עץ בלדרות', index=False, freeze_panes=[1,0],)
                            MishtacheiBaldar_Zikuy_Billing.to_excel(writer, 'זיכוי משטחי עץ בלדרות', index=False, freeze_panes=[1,0],)
                            HafatzaReport_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)
                            ErechMusaf_Billing.to_excel(writer, 'ערך מוסף', index=False, freeze_panes=[1,0],)
                           
                            

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    

                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14
                        ExcelFile['ריכוז'] .column_dimensions['D'].width = 16
                        ExcelFile['ריכוז'] .column_dimensions['E'].width = 16
                        ExcelFile['ריכוז'] .column_dimensions['F'].width = 16

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                       
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      


                        ExcelFile.save(Full_path)

################ גולף סיטונאים  ################################## 
    def GOLFSIT_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select CHARGEID, CHARGELINE, AGREEMENTLINE, CHARGETEXT, BILLFROMDATE, BILLTODATE, BILLTOTAL, UNITS from vBillingChargesByRunIDDetail where CHARGEID={ChargeID} and AGREEMENTLINE in ('30')", connection)
                        Likut_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and AGREEMENTLINE IN ('31','32','33','34') ", connection)
                        Mecholot_Billing = pd.read_sql_query(f"select * from vBillMECHOLOT where CHARGEID='{ChargeID}' ", connection)              
                        Klita_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' AND AGREEMENTLINE IN ('27')", connection)
                        Klita_WTW_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and AGREEMENTLINE='29' ", connection)                        
                        KlitaChazarot_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and AGREEMENTLINE IN ('8','9') ", connection)
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak in ('95')  ", connection)          
                        ErechMusaf_Billing = pd.read_sql_query(f"select * from ProformaSpecialBilling where CHARGEID='{ChargeID}' ", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)                  
                        PickPivotReport_Billing = pd.read_sql_query(f"SELECT CHARGEDESCRIPTION, OBJECTUNITS AS Document_Units, PRICEPERUNIT, VALUE from repProformaDetailed where CHARGEID='{ChargeID}' and CHARGEDESCRIPTION like '%ליקוט%' ", connection)          
                        PickPivotReport_Billing = PickPivotReport_Billing.pivot_table(index='CHARGEDESCRIPTION', values=['Document_Units','PRICEPERUNIT','VALUE'], aggfunc=[sum],margins=True, margins_name='סה"כ')
                        MishtacheiEtzIn_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                         where CHARGEID='{ChargeID}' and AGREEMENTLINE in ('6') ", connection)
                        MishtacheiEtzOut_Billing = pd.read_sql_query(f"select * from repProformaPalltes where CHARGEID='{ChargeID}' ", connection)
                        
                        

                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Likut_Billing.to_excel(writer, 'ליקוטים', index=False, freeze_panes=[1,0],) 
                            PickPivotReport_Billing.to_excel(writer, 'PIVOT ליקוטים', freeze_panes=[1,0],startrow=1)
                            Mecholot_Billing.to_excel(writer, 'מכולות', index=False, freeze_panes=[1,0],)
                            Klita_Billing.to_excel(writer, 'קליטה', index=False, freeze_panes=[1,0],)
                            KlitaChazarot_Billing.to_excel(writer, 'קליטת חזרות', index=False, freeze_panes=[1,0],)
                            MishtacheiEtzIn_Billing.to_excel(writer, 'משטחי עץ-כניסה', index=False, freeze_panes=[1,0],)
                            MishtacheiEtzOut_Billing.to_excel(writer, 'משטחי עץ-יציאה', index=False, freeze_panes=[1,0],) 
                            HafatzaReport_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)
                            Klita_WTW_Billing.to_excel(writer, 'הובלה בהעברה בין מחסנים', index=False, freeze_panes=[1,0],)
                            ErechMusaf_Billing.to_excel(writer, 'ערך מוסף', index=False, freeze_panes=[1,0],)
                           
                            

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    

                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14
                        ExcelFile['ריכוז'] .column_dimensions['D'].width = 16
                        ExcelFile['ריכוז'] .column_dimensions['E'].width = 16
                        ExcelFile['ריכוז'] .column_dimensions['F'].width = 16

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                       
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      


                        ExcelFile.save(Full_path)

################ לבן  ################################## 
    def LAVAN_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select CHARGEID, CHARGELINE, AGREEMENTLINE, CHARGETEXT, BILLFROMDATE, BILLTODATE, BILLTOTAL, UNITS from vBillingChargesByRunIDDetail where CHARGEID={ChargeID} and AGREEMENTLINE in ('6')", connection)
                        Likut_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and CHARGEDESCRIPTION like '%ליקוט%' ", connection)
                        ArizatOnLine_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and AGREEMENTLINE='41' ", connection)
                        Mecholot_Billing = pd.read_sql_query(f"select * from vBillMECHOLOT where CHARGEID='{ChargeID}' and CONSIGNEE='LAVAN'", connection)              
                        Klita_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and AGREEMENTLINE='27' ", connection)
                        Klita_WTW_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and AGREEMENTLINE='29' ", connection)
                        KlitaChazarot_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and AGREEMENTLINE in ('8','9') ", connection)
                        MishtacheiEtz_ZikuyHafatza_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaValueAdded where CHARGEID='{ChargeID}' and AGREEMENTLINE='40' ", connection)
                        MishtacheiEtz_ChiyuvHafatza_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaValueAdded where CHARGEID='{ChargeID}' and AGREEMENTLINE='38' ", connection)
                        #MishtacheiEtzOut_Billing = pd.read_sql_query(f"select * from CheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak in ('53','54') and Teur_Murzar='משטח'  ", connection)
                        MishtacheiBaldar_Chiyuv_Billing = pd.read_sql_query(f"select * from ProformaBaldarutPallets where CHARGEID='{ChargeID}' and LINE='36' ", connection)
                        MishtacheiBaldar_Zikuy_Billing = pd.read_sql_query(f"select * from ProformaBaldarutPallets where CHARGEID='{ChargeID}' and LINE='39' ", connection)
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak='94'  ", connection)          
                        ErechMusaf_Billing = pd.read_sql_query(f"select * from ProformaSpecialBilling where CHARGEID='{ChargeID}' ", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)                  
                        

                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:     
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Likut_Billing.to_excel(writer, 'ליקוטים', index=False, freeze_panes=[1,0],)
                            Mecholot_Billing.to_excel(writer, 'מכולות', index=False, freeze_panes=[1,0],)
                            Klita_Billing.to_excel(writer, 'קליטה', index=False, freeze_panes=[1,0],)
                            KlitaChazarot_Billing.to_excel(writer, 'קליטת חזרות', index=False, freeze_panes=[1,0],)
                            MishtacheiEtz_ChiyuvHafatza_Billing.to_excel(writer, 'משטחי עץ - חיוב הפצה', index=False, freeze_panes=[1,0],)
                            MishtacheiEtz_ZikuyHafatza_Billing.to_excel(writer, 'משטחי עץ - זיכוי הפצה', index=False, freeze_panes=[1,0],)
                            #MishtacheiEtzOut_Billing.to_excel(writer, 'משטחי עץ-יציאה', index=False, freeze_panes=[1,0],)
                            MishtacheiBaldar_Chiyuv_Billing.to_excel(writer, 'חיוב משטחי עץ בלדרות', index=False, freeze_panes=[1,0],)
                            MishtacheiBaldar_Zikuy_Billing.to_excel(writer, 'זיכוי משטחי עץ בלדרות', index=False, freeze_panes=[1,0],)
                            HafatzaReport_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)
                            ErechMusaf_Billing.to_excel(writer, 'ערך מוסף', index=False, freeze_panes=[1,0],)
                            Klita_WTW_Billing.to_excel(writer, 'הובלה בהעברה בין מחסנים', index=False, freeze_panes=[1,0],)
                            ArizatOnLine_Billing.to_excel(writer, 'אריזת הזמנות אונליין', index=False, freeze_panes=[1,0],)

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    

                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                       
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      


                        ExcelFile.save(Full_path)

################ דלתא  ################################## 
    def DELTA_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}'", connection)
                        Mecholot_Billing = pd.read_sql_query(f"select * from vBillMECHOLOT where CHARGEID='{ChargeID}' and CONSIGNEE='DELTA'", connection)              
                        Likut_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and CHARGEDESCRIPTION like '%ליקוט%' ", connection)                        
                        Klita_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTlINE, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and isnull(FIELDVALUE,'')<>'MECHOLOT' ", connection)
                       # MishtacheiEtzIn_Billing = pd.read_sql_query(f"select * from repProformaPalltes where year(DocDate)='{YEARSelected}' and month(DocDate)='{monthSelected}' and BillType='זיכוי' and CONSIGNEE='DELTA' ", connection)
                        MishtacheiEtzIn_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                         where CHARGEID='{ChargeID}' and AGREEMENTLINE in ('6') ", connection)
                        MishtacheiEtzOut_Billing = pd.read_sql_query(f"select * from repProformaPalltes where CHARGEID='{ChargeID}' ", connection)
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak='88'  ", connection)          
                        ErechMusaf_Billing = pd.read_sql_query(f"select * from ProformaSpecialBilling where CHARGEID='{ChargeID}' ", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)                  
                        

                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Mecholot_Billing.to_excel(writer, 'מכולות - פריקה ומיון', index=False, freeze_panes=[1,0],)                            
                            Likut_Billing.to_excel(writer, 'ליקוטים', index=False, freeze_panes=[1,0],)                           
                            Klita_Billing.to_excel(writer, 'קליטה', index=False, freeze_panes=[1,0],)
                            HafatzaReport_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)
                            MishtacheiEtzIn_Billing.to_excel(writer, 'משטחי עץ לחיוב', index=False, freeze_panes=[1,0],)
                            MishtacheiEtzOut_Billing.to_excel(writer, 'משטחי עץ לזיכוי', index=False, freeze_panes=[1,0],)                            
                            ErechMusaf_Billing.to_excel(writer, 'ערך מוסף מיון ובניה', index=False, freeze_panes=[1,0],)
                            

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    

                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                       
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      


                        ExcelFile.save(Full_path)

################ גולף  ##################################
    def GOLF_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}'", connection)
                        Mecholot_Billing = pd.read_sql_query(f"select * from vBillMECHOLOT where CHARGEID='{ChargeID}' and CONSIGNEE='GOLF'", connection)              
                        Likut_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and CHARGEDESCRIPTION like '%ליקוט%' ", connection)                        
                        Klita_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT  \
                                                                    , CASE WHEN TRANSPORTTYPE='PurchaseOrder' THEN OBJECTUNITS ELSE 0 END AS PalltesToBill from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and isnull(FIELDVALUE,'')<>'MECHOLOT' and AGREEMENTLINE=4 ", connection)
                        HovalaFromGolf = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and isnull(FIELDVALUE,'')<>'MECHOLOT' and AGREEMENTLINE in ('5','6') ", connection)
                        HovalaToGolf = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and AGREEMENTLINE='3' ", connection)                 
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)                  
                        MishtachimLezikuy_Billing = pd.read_sql_query(f"select sum(CAST(qty as int)) from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak=2 and Teur_Godel_Mechiron='משטח-גובני'  ", connection)          
                        ZikuyMishtacheiEtz_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaValueAdded where CHARGEID='{ChargeID}' and AGREEMENTLINE='11' ", connection)
                        ChiyuvMishtacheiEtzHafatza_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT  \
                                                                    , CASE WHEN TRANSPORTTYPE='PurchaseOrder' THEN OBJECTUNITS ELSE 0 END AS PalltesToBill from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and AGREEMENTLINE in ('12','13','14') ", connection)
                        Asortiment_Billing = pd.read_sql_query(f"select * from repProformaWcbOXESLTR where CHARGEID='{ChargeID}' and AGREEMENTLINE in ('6','7','8')", connection)
                        

                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Mecholot_Billing.to_excel(writer, 'מכולות - פריקה ומיון', index=False, freeze_panes=[1,0],)                            
                            Likut_Billing.to_excel(writer, 'ליקוטים', index=False, freeze_panes=[1,0],)                           
                            Klita_Billing.to_excel(writer, 'קליטה', index=False, freeze_panes=[1,0],)
                            HovalaToGolf.to_excel(writer, 'הובלה לגולף', index=False, freeze_panes=[1,0],)
                            HovalaFromGolf.to_excel(writer, 'הובלה מגולף', index=False, freeze_panes=[1,0],)
                            ChiyuvMishtacheiEtzHafatza_Billing.to_excel(writer,'משטחי עץ - חיוב', index=False, freeze_panes=[1,0],)
                            ZikuyMishtacheiEtz_Billing.to_excel(writer,'משטחי עץ - זיכוי', index=False, freeze_panes=[1,0],)
                            Asortiment_Billing.to_excel(writer, 'ערך מוסף אסורטימנטים LTR', index=False, freeze_panes=[1,0],)

                           
                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    

                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14
                        ExcelFile['ריכוז'] .column_dimensions['G'].width = 37
                        ExcelFile['ריכוז'] .column_dimensions['E'].width = 16
                        ExcelFile['ריכוז'] .column_dimensions['D'].width = 16
                        ExcelFile['ריכוז'] .column_dimensions['F'].width = 22
                        #ExcelFile['ריכוז'] ['F12'] = MishtachimLezikuy_Billing.iloc[0][0]
                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                       
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      


                        ExcelFile.save(Full_path)

################ ג'ינגר  ################################## 
    def JINGER_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}'", connection)
                        loads_billing.drop(['UNITS'], axis=1, inplace=True)
                        Likut_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and CHARGEDESCRIPTION like '%ליקוט%' ", connection)
                        Mecholot_Billing = pd.read_sql_query(f"select * from vBillMECHOLOT where CHARGEID='{ChargeID}' and CONSIGNEE='JIN'", connection)              
                        Klita_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and AGREEMENTLINE in ('12') ", connection)
                        KlitaChazarot_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and isnull(FIELDVALUE,'')<>'MECHOLOT' and AGREEMENTLINE in ('5','6') ", connection)
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak='83'  ", connection)          
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)                  
                    
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Mecholot_Billing.to_excel(writer, 'מכולות', index=False, freeze_panes=[1,0],)
                            Klita_Billing.to_excel(writer, 'קליטה', index=False, freeze_panes=[1,0],)
                            Likut_Billing.to_excel(writer, 'ליקוטים', index=False, freeze_panes=[1,0],)
                            KlitaChazarot_Billing.to_excel(writer, 'קליטת החזרות', index=False, freeze_panes=[1,0],)
                            HafatzaReport_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)                            

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    
                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14
                        ExcelFile['ריכוז'] .column_dimensions['C'].width = 10

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                       
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      


                        ExcelFile.save(Full_path)

################ יקב מורד  ################################## 
    def MORAD_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        loads_billing = pd.read_sql_query(f"select * from repProformaLoads where CHARGEID='{ChargeID}'", connection)
                        Likut_Billing = pd.read_sql_query(f"select CHARGEID, AGREEMENTNAME, AGREEMENTLINE, CHARGEDESCRIPTION, AGREEMENTDESC, CONSIGNEENAME, ROW_NUMBER() OVER(ORDER BY CHARGEID) AS LineNumber \
                                    , COMPANY, COMPANYNAME, Contact, CONTACT1NAME, OBJECTID AS Document, OBJECTDATE AS Document_Date, OBJECTUNITS AS Document_Units \
                                    , PRICEPERUNIT, VALUE, SHIPPEDDATE, REFERENCEORD AS OutboundReference, ParamREFERENCEORDER AS OutParamReference \
                                    , SECAGENT, AGENTDESC, ORDERTYPE from repProformaDetailed where CHARGEID='{ChargeID}' and AGREEMENTLINE in ('6','3') ", connection)
                        Klita_Billing = pd.read_sql_query(f"select CHARGEID,CONSIGNEE, CHARGELINE,AGREEMENTNAME, AGREEMENTLINE, COMPANYNAME, OBJECTID AS DOCUMENT, BOL,CONVERT(DATE, OBJECTDATE) AS DocumentDate \
                                                                    , OBJECTUNITS as UNITS, PRICEPERUNIT, VALUE as CHAREGE, CLOSERECEIPTDATE as DATE, VEHICLE as VENDOR, DRIVER1 as CONTACT from repProformaInDetailed \
                                                                        where CHARGEID='{ChargeID}' and  AGREEMENTLINE in ('1','2')", connection)
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak='37'  ", connection)          
                        Rikuz_Billing = pd.read_sql_query(f"select CONSIGNEENAME from CONSIGNEE where CONSIGNEE='{consigneeSelected}'", connection)                  
                    
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)
                            loads_billing.to_excel(writer,'אחסנה',index=False, freeze_panes=[1,0],)
                            Klita_Billing.to_excel(writer, 'קליטה', index=False, freeze_panes=[1,0],)
                            Likut_Billing.to_excel(writer, 'ליקוטים והמכלות', index=False, freeze_panes=[1,0],)
                            HafatzaReport_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)                            

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    
                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                       
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      


                        ExcelFile.save(Full_path)

################ קרביץ  ################################## 
    def KRAVITZ_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak IN ('1900','19')  ", connection)          
                        Rikuz_Billing = pd.read_sql_query(f"select 'קרביץ' ", connection)                  
                    
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)                           
                            HafatzaReport_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)                            

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    
                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                       
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      


                        ExcelFile.save(Full_path)

################ גולף CD ################################## 
    def GOLF_CD_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak IN ('2')  ", connection)          
                        Rikuz_Billing = pd.read_sql_query(f"select 'GOLF_CD' ", connection)                  
                    
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)                           
                            HafatzaReport_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)                            

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    
                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                       
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      


                        ExcelFile.save(Full_path)

################ אשד ################################## 
    def ESHED_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak IN ('3')  ", connection)          
                        Rikuz_Billing = pd.read_sql_query(f"select 'ESHED' ", connection)                  
                    
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)                           
                            HafatzaReport_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)                            

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    
                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                       
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      


                        ExcelFile.save(Full_path)

################ KNS ################################## 
    def KNS_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak IN ('69')  ", connection)          
                        Rikuz_Billing = pd.read_sql_query(f"select 'KNS' ", connection)                  
                    
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)                           
                            HafatzaReport_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)                            

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    
                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                       
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      


                        ExcelFile.save(Full_path)

################ לוגיסטים ################################## 
    def LOGISTEAM_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        HafatzaReport_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak IN ('101','150')  ", connection)          
                        Rikuz_Billing = pd.read_sql_query(f"select 'לוגיסטים'", connection)                  
                    
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)                           
                            HafatzaReport_Billing.to_excel(writer, 'הפצה', index=False, freeze_panes=[1,0],)                            

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    
                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14
                        ExcelFile['ריכוז'] .column_dimensions['C'].width = 10
                        ExcelFile['ריכוז'] .column_dimensions['D'].width = 20

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי                     
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'   

                         # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='C'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי                     
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      
   


                        ExcelFile.save(Full_path)

################ טסט ################################## 
    def TEST_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        HafatzaReport_Billing = pd.read_sql_query(f"SELECT * from repProformaDetailed where CHARGEID='0000005607' and CHARGEDESCRIPTION like '%ליקוט%' ", connection)          
                        HafatzaPivotReport_Billing = pd.read_sql_query(f"SELECT CHARGEDESCRIPTION, OBJECTUNITS AS Document_Units, PRICEPERUNIT, VALUE from repProformaDetailed where CHARGEID='0000005607' and CHARGEDESCRIPTION like '%ליקוט%' ", connection)          
                        HafatzaPivotReport_Billing = HafatzaPivotReport_Billing.pivot_table(index='CHARGEDESCRIPTION', values=['Document_Units','PRICEPERUNIT','VALUE'], aggfunc=[sum],margins=True, margins_name='סה"כ')
                        Rikuz_Billing = pd.read_sql_query(f"select 'TEST' ", connection)                  
                    
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:  
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)                              
                            HafatzaReport_Billing.to_excel(writer, 'טסט', index=False, freeze_panes=[1,0],)
                            HafatzaPivotReport_Billing.to_excel(writer, 'פיווט', freeze_panes=[1,0],startrow=1)  
                                                        

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    
                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי                     
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'   

                         # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='C'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי                     
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      
   


                        ExcelFile.save(Full_path)

################ אורשר  ################################## 
    def ORSHAR_EXCEL(consigneeSelected, Full_path, YEARSelected, monthSelected, ChargeID):
                        #שליפת הנתונים ללשוניות הרלוונטיות
                        HafatzaReport_NirGalim_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak IN ('41')  ", connection)
                        HafatzaReport_Tornado_Billing = pd.read_sql_query(f"select * from vCheshbonSapakimMegicBill where year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and Mispar_Sapak IN ('48')  ", connection)                    
                        Hafatza_Rikuz_NirGalim_Billing = pd.read_sql_query(f"select * from vPivotCheshbonSapakimMegicBill where  year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and CONSIGNEE='ORSHAR_NIRGALIM' ", connection)
                        Hafatza_Rikuz_Tornado_Billing = pd.read_sql_query(f"select * from vPivotCheshbonSapakimMegicBill where  year(Date_Aspaka)='{YEARSelected}' and month(Date_Aspaka)='{monthSelected}' and CONSIGNEE='ORSHAR_TORNADO' ", connection)
                        Rikuz_Billing = pd.read_sql_query(f"select 'אורשר' ", connection)                  
                    
                        #יצירת קובץ האקסל
                        with pd.ExcelWriter(Full_path) as writer:          
                            Rikuz_Billing.to_excel(writer,'ריכוז',index=False, freeze_panes=[1,0],header=False)                           
                            HafatzaReport_NirGalim_Billing.to_excel(writer, 'פירוט ניר גלים', index=False, freeze_panes=[1,0],)                            
                            HafatzaReport_Tornado_Billing.to_excel(writer, 'פירוט טורנדו', index=False, freeze_panes=[1,0],)
                            Hafatza_Rikuz_NirGalim_Billing.to_excel(writer, 'ריכוז ניר גלים', index=False, freeze_panes=[1,0],)                            
                            Hafatza_Rikuz_Tornado_Billing.to_excel(writer, 'ריכוז טורנדו', index=False, freeze_panes=[1,0],)                                                        

                        # הפעלת פונקציה לעיצוב האקסל
                        ExcelDesign(Full_path)
                        #שמירת הקובץ במשתנה לשימוש בהמשך הקוד
                        ExcelFile = load_workbook(Full_path)
                    
                          ## בניית תוכן של לשונית ריכוז 
                        ExcelBillAppCellValue = pd.read_sql_query(f"select * from ExcelBillAppCellValue where consignee='{consigneeSelected}'", connection)               

                        #השמת ערכים בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Update = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Value']
                        for row in Excel_Value_Update.index:
                            ExcelFile[ExcelBillAppCellValue['TAB'][row]] [ExcelBillAppCellValue['CELL'][row]] = ExcelBillAppCellValue['Cell_Value'][row]
                       
                        #עדכון מילוי בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Fill = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Fill']                       
                        for row in Excel_Value_Fill.index:
                            ExcelFile[Excel_Value_Fill['TAB'][row]][Excel_Value_Fill['CELL'][row]].fill = Choose_Fill(Excel_Value_Fill['Cell_Value'][row])
                       
                        #עדכון פונט בתאים לפי שליפה מערכים בבסיס הנתונים
                        Excel_Value_Font = ExcelBillAppCellValue[ExcelBillAppCellValue['Change_Type'] == 'Font']                       
                        for row in Excel_Value_Font.index:
                            ExcelFile[Excel_Value_Font['TAB'][row]][Excel_Value_Font['CELL'][row]].font = Choose_Font(Excel_Value_Font['Cell_Value'][row])                              

                        # ExcelFile['ריכוז'] ['B3'].fill = Header_fill
                        # ExcelFile['ריכוז'] ['A1'].font = Header_font
                        # ExcelFile['ריכוז'] ['A4'] = 'אחסנה'
                        # ExcelFile['ריכוז'] ['B4'] = '=SUM(אחסנה!V:V)'
                        
                        #עדכון מיזוג וגודל עמודות קבוע ללשונית ריכוז 
                        ExcelFile['ריכוז'] .merge_cells('A1:B1')
                        ExcelFile['ריכוז'] .column_dimensions['A'].width = 22
                        ExcelFile['ריכוז'] .column_dimensions['B'].width = 14

                        # להפוך את כל התאים של מחיר לפורמט שקל
                        Charge_cell = ExcelBillAppCellValue[ExcelBillAppCellValue['CELL'].astype(str).str[0] =='B'] #שליפה של השורות שהערך בעמודה תא מתחיל האות בי

                       
                        for row in Charge_cell.index:                           
                            ExcelFile['ריכוז'] [Charge_cell['CELL'][row]].number_format = u'#,##0 ₪'      


                        ExcelFile.save(Full_path)

    ###############################################################################################
    ####################קוד רץ של העמוד###########################################################
    with st.form('MyFirstForm'):  
            c1, c2, c3 = st.columns(3)   
            with c1:      
                consigneeSelected = st.selectbox('Consignee', options=consignee)
            with c2:
                monthSelected = st.selectbox('month',['01','02','03','04','05','06','07','08','09','10','11','12'])
            with c3:
                YEARSelected = st.selectbox('year', options=YearOptions)
            path = st.text_input('Folder path')               
            File_name = st.text_input('File name')
            Full_path = path + '\\' + consigneeSelected + '-' +  YEARSelected.strip() + '-' + monthSelected  + File_name + '.xlsx'
            

            Submit = st.form_submit_button('Create billing report')      
            if Submit:
                #בדיקה על השדות שהתמלאו בצורה תקינה
                if path == '':
                    st.error('Folder field is empty')              
                if not(os.path.exists(path)):
                    st.error('Choosen folder path not exists')
                    
                else:
                #  שליפת מספר חיוב רלוונטי לחודש למאחסן בשביל שליפת המידע ללשוניות אחסנה יציאות וכניסות כולל בדיקה שיש רק חיוב אחד פתוח לאותון חודש למאחסן 
                    NumChargeInMonth = pd.read_sql_query(f"select count(*) as NumCharges from BILLINGCHARGESHEADER where consignee='{consigneeSelected}' and year(BILLTODATE)='{YEARSelected}' and month(BILLTODATE)='{monthSelected}' and STATUS='NEW' ", connection).iloc[0]['NumCharges']
                    if NumChargeInMonth>1:  
                        st.error('יש יותר מחיוב אחד חדש עבור המאחסן לחודש שנבחר')
                        st.stop()
                    ChargeID = pd.read_sql_query(f"select isnull(max(CHARGEID),'') as CHARGEID from BILLINGCHARGESHEADER where consignee='{consigneeSelected}' and year(BILLTODATE)='{YEARSelected}' and month(BILLTODATE)='{monthSelected}'", connection).iloc[0]['CHARGEID'].replace(" ", "")
                
                    if consigneeSelected=='ELECTRA':
                        ELECTRA_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='BUR':
                        BUR_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='BIAPAL':
                        BFL_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='AFRODITA':
                        AFRODITA_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='BONITA':
                        BONITA_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='KITAN':
                        KITAN_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='LAVAN':
                        LAVAN_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='JIN':
                        JINGER_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='MORAD':
                        MORAD_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='DELTA':
                        DELTA_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='GOLF':
                        GOLF_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='TENELECTRIK':
                        TENELECTRIK_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='MANIA':
                        MANIA_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='TADIRANG':
                        TADIRAN_GROUP_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='FLEX':
                        FLEX_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='KRAVITZ':
                        KRAVITZ_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='ORSHAR':
                        ORSHAR_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='GOLF_CD':
                        GOLF_CD_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='ESHED':
                        ESHED_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='KNS':
                        KNS_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='LOGISTEAM':
                        LOGISTEAM_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='TEST':
                        TEST_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='MCHASHMAL':
                        MCHASHMAL_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='SHEROTHOGEN':
                        SHEROTHOGEN_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    elif consigneeSelected=='GOLFSIT':
                        GOLFSIT_EXCEL(consigneeSelected,Full_path,YEARSelected, monthSelected, ChargeID)
                    else:
                        st.error(f'File not created')
                        st.stop()

                    st.write(f'ChargeId Selected: {ChargeID}')
                    st.success('File created successfully')              
                    st.write('File path is:   \n'  + Full_path)
                



elif choose == "Report Design":
    PageName = '<p style="font-family:sans-serif; color:black; font-size: 20px; background: #F0F2F6; text-align: center; ">Report Design</p>'
    st.markdown(PageName, unsafe_allow_html=True)


elif choose == "Wms changes":
    PageName = '<p style="font-family:sans-serif; color:black; font-size: 20px; background: #F0F2F6; text-align: center; ">Wms changes</p>'
    st.markdown(PageName, unsafe_allow_html=True)


elif choose == "Compare Report":
    PageName = '<p style="font-family:sans-serif; color:black; font-size: 20px; background: #F0F2F6; text-align: center; ">Compare Report</p>'
    st.markdown(PageName, unsafe_allow_html=True)
